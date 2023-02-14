import openpyxl 
from openpyxl.utils.exceptions import InvalidFileException

def get_matrix_of_products() -> object:
    try:
        SHEET_PATH = 'C:\\Users\\conta\\OneDrive\\Documentos\\scarperdeestoqueortoestetica\\estoqueoeg.xlsx'
        matrix_of_items = []

        work_book = openpyxl.load_workbook(SHEET_PATH)
        sheet_object = work_book.active 
        m_row = sheet_object.max_row

        for i in range(1, m_row + 1):
            items = []
            materials_name = sheet_object.cell(row = i, column = 1) 
            description = 'Quantidade em estoque - '+ str(sheet_object.cell(row = i, column = 2).value)
            measurement_unit = sheet_object.cell(row = i, column = 3)     
            items.append(materials_name.value)
            items.append(measurement_unit.value)
            items.append(description)
            matrix_of_items.append(items)

        return matrix_of_items
    except InvalidFileException:
        raise InvalidFileException('Could not read the file!')